from pdfminer.high_level import extract_pages
from matplotlib import pyplot as plt
from pdf2image import convert_from_path
from os import listdir
from matplotlib.widgets import RectangleSelector
from camelot import read_pdf
from numpy import abs
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import read_excel
from datetime import date
from os import path


def mostrar_pagina(arquivo, numero_pagina):
    # Variáveis importantes para a função mouse_click()
    global rs, x_coords, y_coords
    x_coords = []
    y_coords = []

    def mouse_click(event):
        global x_coords, y_coords
        x, y = event.xdata, event.ydata
        
        x_coords = [x]
        y_coords = [y]

    def mouse_release(event):
        global x_coords, y_coords

        x, y = event.xdata, event.ydata

        x_coords.append(x)
        y_coords.append(y)

        coords = f'{min(x_coords)},{max(y_coords)},{max(x_coords)},{min(y_coords)}'

        table = read_pdf(arquivo, pages=str(numero_pagina+1), flavor='stream', 
                                    table_areas=[coords], flag_size=False)
        print("Total tables extracted:", table.n)
        df = table[0].df
        print(df)
        print("Tabela enviada para Excel")
        #df.to_excel("Teste.xlsx")
        write_excel(df, arquivo)

    # Poppler e foto da página
    poppler = "C:/Program Files/poppler-22.01.0/Library/bin"
    image = convert_from_path(arquivo, poppler_path=poppler, first_page=numero_pagina+1, last_page=numero_pagina+1)

    # Layout da página (dimensões e elementos)
    pages = extract_pages(arquivo, page_numbers=[numero_pagina+1])

    page_layouts = []
    for page_layout in pages:
        page_layouts.append(page_layout)
    
    # Plotar a figura com as dimensões corretas
    x1 = page_layouts[0].x1
    y1 = page_layouts[0].y1

    fig, ax = plt.subplots()
    ax.imshow(image[0], extent=[0, x1, 0, y1])

    def line_select_callback(eclick, erelease):
        x1, y1 = eclick.xdata, eclick.ydata
        x2, y2 = erelease.xdata, erelease.ydata

        rect = plt.Rectangle((min(x1,x2),min(y1,y2)), abs(x1-x2), abs(y1-y2))
        ax.add_patch(rect)

    rs = RectangleSelector(ax, line_select_callback,
                       drawtype='box', useblit=True, button=[1], 
                       minspanx=5, minspany=5, spancoords='pixels', 
                       interactive=False, rectprops=dict(facecolor='blue', edgecolor='blue', alpha=0.2, fill=True))

    plt.connect('button_press_event', mouse_click)
    plt.connect('button_release_event', mouse_release)
    plt.show()


def apresentar_arquivos():
    # Lista todos os pdfs na pasta
    files = list()
    lista = listdir("Examples")
    for file in lista:
        if '.pdf' in file:
            files.append(file)
        else:
            pass

    return files



def write_excel(df, arquivo):
    today = date.today()
    dest_filename = f'Sheets/{today.day}-{today.month}-{today.year}.xlsx'
    excel_exists = path.exists(dest_filename)
    sheet_name = arquivo.split('/')[-1]

    if excel_exists:
        wb = load_workbook(filename=dest_filename)

        if sheet_name in wb.sheetnames:
            sheet_data = read_excel(dest_filename, sheet_name=sheet_name)
            lines_used = len(sheet_data) + 5
            ws = wb[sheet_name]
        else:
            lines_used = 0
            ws = wb.create_sheet(sheet_name)
    
    else:
        wb = Workbook()
        lines_used = 0
        ws = wb.create_sheet(sheet_name)


    rows = dataframe_to_rows(df)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx+lines_used, column=c_idx, value=value)


    wb.save(filename = dest_filename)