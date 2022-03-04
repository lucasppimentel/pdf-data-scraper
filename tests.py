from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame, read_excel
from datetime import date
from os import path

# This code works, need to make it into a function

today = date.today()
dest_filename = f'Sheets/{today.day}-{today.month}-{today.year}.xlsx'
sheet_exists = path.exists(dest_filename)

if sheet_exists:
    wb = load_workbook(filename=dest_filename)
else:
    wb = Workbook()

df = read_excel(dest_filename, sheet_name="Sheet") # Change name of sheet to the name of the pdf file
lines_used = len(df) + 5

db = DataFrame(data=[[1, 2, 3],[4,5,6]]) # Just a test, that's the data frame I wanna write

ws = wb["Sheet"] # Change name of sheet to the name of the pdf file


rows = dataframe_to_rows(db)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
         ws.cell(row=r_idx+lines_used, column=c_idx, value=value)


wb.save(filename = dest_filename)